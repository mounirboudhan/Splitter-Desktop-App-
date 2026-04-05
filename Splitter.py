import os
import sys
import ctypes
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import win32com.client
import threading
import re
import pythoncom
import random
import string
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import Workbook, load_workbook

# -------------------------------------------------------------
# Project Name : Splitter - Document into individually named PDF pages
# Developed By : Mounir Boudhan
# Description  : Convert Word doc into individual PDFs per page.
# -------------------------------------------------------------

# DPI Awareness for high resolution screens
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

# UI Theme Configuration: Modern Light Style
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class SplitterApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Splitter - Document into individually named PDF pages")
        self.geometry("720x620")
        self.resizable(False, False)

        # Set window icon
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(base_path, "favicon.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)
        
        # Color palette - Modern Blue Theme
        self.bg_color = "#F8FAFC"        # Slate 50
        self.card_bg = "#FFFFFF"         # Pure White
        self.accent_primary = "#2563EB"  # True Blue
        self.accent_secondary = "#1D4ED8" # Darker Blue
        self.accent_success = "#2563EB"   # True Blue
        self.text_main = "#1E293B"        # Slate 800
        self.text_muted = "#64748B"       # Slate 500
        self.border_clr = "#E2E8F0"       # Slate 200
        self.input_bg = "#FFFFFF"         # White for inputs

        self.configure(fg_color=self.bg_color)

        # Center window on screen
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f'+{x}+{y}')

        # Main Layout
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Container
        self.container = ctk.CTkFrame(
            self, fg_color=self.card_bg, corner_radius=20,
            border_width=1, border_color=self.border_clr
        )
        self.container.grid(padx=30, pady=20, sticky="nsew")
        self.container.grid_columnconfigure(0, weight=1)

        # Header
        self.header_frame = ctk.CTkFrame(self.container, fg_color="transparent")
        self.header_frame.grid(row=0, column=0, pady=(20, 15), sticky="ew")
        
        self.title_label = ctk.CTkLabel(
            self.header_frame, text="Document Splitter",
            font=ctk.CTkFont(size=30, weight="bold"),
            text_color=self.text_main
        )
        self.title_label.pack()
        
        self.subtitle_label = ctk.CTkLabel(
            self.header_frame, text="Split Word documents into individually named PDF pages.",
            font=ctk.CTkFont(size=14), text_color=self.text_muted
        )
        self.subtitle_label.pack(pady=(2, 0))

        # Content
        self.content_frame = ctk.CTkFrame(self.container, fg_color="transparent")
        self.content_frame.grid(row=1, column=0, padx=60, sticky="ew")
        self.content_frame.grid_columnconfigure(0, weight=1)

        # Step 1
        self.input_label = ctk.CTkLabel(
            self.content_frame, text="STEP 1: SOURCE DOCUMENT",
            font=ctk.CTkFont(size=11, weight="bold"), text_color=self.accent_primary
        )
        self.input_label.grid(row=0, column=0, sticky="w", pady=(0, 5))

        self.input_entry = ctk.CTkEntry(
            self.content_frame, placeholder_text="Choose .docx...",
            height=40, font=ctk.CTkFont(size=12), fg_color=self.input_bg, border_color=self.border_clr, text_color=self.text_main
        )
        self.input_entry.insert(0, "No document selected...")
        self.input_entry.configure(state="disabled")
        self.input_entry.grid(row=1, column=0, sticky="ew", padx=(0, 10))

        self.input_button = ctk.CTkButton(
            self.content_frame, text="Browse", command=self.browse_file,
            width=100, height=40, fg_color=self.accent_primary,
            hover_color=self.accent_secondary, font=ctk.CTkFont(weight="bold")
        )
        self.input_button.grid(row=1, column=1)

        # Step 2: Passwords Excel (Optional)
        self.excel_label = ctk.CTkLabel(
            self.content_frame, text="STEP 2: PASSWORDS EXCEL (OPTIONAL)",
            font=ctk.CTkFont(size=11, weight="bold"), text_color=self.accent_primary
        )
        self.excel_label.grid(row=2, column=0, sticky="w", pady=(10, 5))

        self.excel_entry = ctk.CTkEntry(
            self.content_frame, placeholder_text="Excel with ID / Password columns...",
            height=40, font=ctk.CTkFont(size=12), fg_color=self.input_bg, border_color=self.border_clr, text_color=self.text_main
        )
        self.excel_entry.insert(0, "No Excel map selected (Optional)...")
        self.excel_entry.configure(state="disabled")
        self.excel_entry.grid(row=3, column=0, sticky="ew", padx=(0, 10))

        self.excel_button = ctk.CTkButton(
            self.content_frame, text="Browse", command=self.browse_excel,
            width=100, height=40, fg_color="#475569",
            hover_color="#334155", font=ctk.CTkFont(weight="bold")
        )
        self.excel_button.grid(row=3, column=1)

        # Step 3: Output Directory
        self.output_label = ctk.CTkLabel(
            self.content_frame, text="STEP 3: OUTPUT DIRECTORY",
            font=ctk.CTkFont(size=11, weight="bold"), text_color=self.accent_primary
        )
        self.output_label.grid(row=4, column=0, sticky="w", pady=(10, 5))

        self.output_entry = ctk.CTkEntry(
            self.content_frame, placeholder_text="Saving to...",
            height=40, font=ctk.CTkFont(size=12), fg_color=self.input_bg, border_color=self.border_clr, text_color=self.text_main
        )
        self.output_entry.insert(0, "No output directory selected...")
        self.output_entry.configure(state="disabled")
        self.output_entry.grid(row=5, column=0, sticky="ew", padx=(0, 10))

        self.output_button = ctk.CTkButton(
            self.content_frame, text="Choose", command=self.browse_folder,
            width=100, height=40, fg_color="#475569",
            hover_color="#334155", font=ctk.CTkFont(weight="bold")
        )
        self.output_button.grid(row=5, column=1)

        # Status Area
        self.status_frame = ctk.CTkFrame(self.container, fg_color="#F1F5F9", corner_radius=10, border_width=1, border_color=self.border_clr)
        self.status_frame.grid(row=2, column=0, padx=60, pady=15, sticky="ew")
        self.status_frame.grid_columnconfigure(0, weight=1)
        
        self.status_label = ctk.CTkLabel(
            self.status_frame, text="Ready to split",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=self.text_muted
        )
        self.status_label.grid(row=0, column=0, pady=(10, 5), sticky="ew")

        self.progress_bar = ctk.CTkProgressBar(
            self.status_frame, height=6,
            progress_color=self.accent_primary, fg_color="#E2E8F0"
        )
        self.progress_bar.set(0)
        self.progress_bar.grid(row=1, column=0, pady=(0, 8), padx=20, sticky="ew")

        # View Results Button — always in row=2, just hidden via grid_remove
        self.view_results_btn = ctk.CTkButton(
            self.status_frame, text="VIEW GENERATED FILES",
            command=self.open_output_folder, height=28, width=200,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#64748B", hover_color="#475569"
        )
        # Reserve the row but keep it invisible until needed
        self.view_results_btn.grid(row=2, column=0, pady=(0, 10))
        self.view_results_btn.grid_remove()   # hidden but space NOT reserved

        # Main Action Button
        self.process_button = ctk.CTkButton(
            self.container, text="START CONVERSION",
            command=self.handle_process_click,
            height=45, font=ctk.CTkFont(size=14, weight="bold"),
            corner_radius=8, fg_color=self.accent_success, hover_color=self.accent_secondary
        )
        self.process_button.grid(row=3, column=0, pady=(0, 20), padx=60, sticky="ew")

        # Footer
        self.footer_label = ctk.CTkLabel(
            self.container, text="v1.1.0 • Developed by Mounir Boudhan",
            font=ctk.CTkFont(size=10), text_color=self.text_muted
        )
        self.footer_label.grid(row=4, column=0, pady=(0, 10))

        # State
        self.is_processing = False
        self.stop_requested = False
        self.current_out_dir = ""

    # ─── File/Folder Browsing ───────────────────────────────────
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Select Word Document",
            filetypes=[("Word Documents", "*.docx")]
        )
        if filename:
            self.input_entry.configure(state="normal")
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, filename)
            self.input_entry.configure(state="disabled")
            if not self.output_entry.get():
                self.output_entry.configure(state="normal")
                self.output_entry.insert(0, os.path.dirname(filename))
                self.output_entry.configure(state="disabled")

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select Output Directory")
        if folder:
            self.output_entry.configure(state="normal")
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, folder)
            self.output_entry.configure(state="disabled")

    def browse_excel(self):
        filename = filedialog.askopenfilename(
            title="Select Passwords Excel Mapping",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if filename:
            self.excel_entry.configure(state="normal")
            self.excel_entry.delete(0, tk.END)
            self.excel_entry.insert(0, filename)
            self.excel_entry.configure(state="disabled")

    # ─── Thread-safe UI helpers ─────────────────────────────────
    def _set_status(self, text, progress=None):
        """Must be called from main thread only (via self.after)."""
        self.status_label.configure(text=text)
        if progress is not None:
            self.progress_bar.set(progress)

    def update_status(self, text, progress=None):
        """Safe to call from any thread."""
        self.after(0, lambda: self._set_status(text, progress))

    def _show_results_btn(self):
        self.view_results_btn.grid()       # restore in its fixed row

    def _hide_results_btn(self):
        self.view_results_btn.grid_remove()  # hide without shifting layout

    def _reset_button(self):
        self.process_button.configure(
            state="normal", text="START CONVERSION",
            fg_color=self.accent_primary, hover_color=self.accent_secondary
        )

    def _reset_for_new_run(self):
        """Fully reset UI so the user can start a new conversion immediately."""
        self._reset_button()

    def _set_button_stopping(self):
        self.process_button.configure(state="disabled", text="CANCELING...")

    def _set_button_running(self):
        self.process_button.configure(
            text="⬛ STOP PROCESS", fg_color="#EF4444", hover_color="#DC2626"
        )

    # ─── Process Control ────────────────────────────────────────
    def handle_process_click(self):
        if self.is_processing:
            # User wants to stop
            self.stop_requested = True
            self.after(0, self._set_button_stopping)
        else:
            # User wants to start
            self.start_conversion()

    def open_output_folder(self):
        if self.current_out_dir and os.path.isdir(self.current_out_dir):
            os.startfile(self.current_out_dir)

    def start_conversion(self):
        docx_path = self.input_entry.get()
        out_dir = self.output_entry.get()
        excel_mapping_path = self.excel_entry.get()

        if not docx_path or not os.path.exists(docx_path):
            messagebox.showerror("Error", "Please select a valid Word document.")
            return
        if not out_dir or not os.path.isdir(out_dir):
            messagebox.showerror("Error", "Please select a valid output directory.")
            return

        self.current_out_dir = out_dir
        self.is_processing = True
        self.stop_requested = False
        self.after(0, self._hide_results_btn)
        self.after(0, self._set_button_running)

        thread = threading.Thread(
            target=self.run_conversion, args=(docx_path, out_dir, excel_mapping_path), daemon=True
        )
        thread.start()

    # ─── Core Conversion (runs in background thread) ────────────
    def run_conversion(self, docx_path, out_dir, excel_mapping_path):
        word_app = None
        doc = None
        try:
            # ── Load Password Mapping (if provided) ──
            id_to_password_map = {}
            if excel_mapping_path and os.path.exists(excel_mapping_path):
                self.update_status("Loading password mapping...", 0.02)
                try:
                    wb_map = load_workbook(excel_mapping_path, data_only=True)
                    ws_map = wb_map.active
                    # Assume headers "ID" and "Password" or just Col A=ID, Col B=Password
                    # We'll check the first row for headers
                    col_id = 0
                    col_pwd = 1  # Default to the second column as requested
                    start_row = 1

                    first_row = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in ws_map[1]]
                    
                    if "id" in first_row:
                        col_id = first_row.index("id")
                    if "password" in first_row:
                        col_pwd = first_row.index("password")
                    elif "pass" in first_row:
                        col_pwd = first_row.index("pass")
                    
                    # If we found headers, start from row 2
                    if "id" in first_row or "password" in first_row or "pass" in first_row:
                        start_row = 2
                    
                    for row in ws_map.iter_rows(min_row=start_row, values_only=True):
                        if row and len(row) > max(col_id, col_pwd):
                            m_id = str(row[col_id]).strip()
                            m_pwd = str(row[col_pwd]).strip()
                            
                            if m_id and m_pwd and m_id != "None" and m_pwd != "None":
                                # Handle numeric IDs that openpyxl reads as "1234.0"
                                if m_id.endswith(".0"):
                                    m_id = m_id[:-2]
                                    
                                # Normalize ID for comparison (remove spaces, slashes, dashes)
                                clean_m_id = re.sub(r'[\s/-]', '', m_id)
                                id_to_password_map[clean_m_id] = m_pwd
                except Exception as ex:
                    print(f"Error loading Excel mapping: {ex}")

            self.update_status("Initializing Word Engine...", 0.05)
            pythoncom.CoInitialize()
            
            word_app = win32com.client.DispatchEx("Word.Application")
            word_app.Visible = False
            word_app.DisplayAlerts = 0  # wdAlertsNone - suppress ALL dialogs
            word_app.AutomationSecurity = 3  # msoAutomationSecurityForceDisable - no macros
            
            self.update_status("Opening document...", 0.1)
            # Normalize path to Windows backslash format for COM
            abs_docx = os.path.normpath(os.path.abspath(docx_path))
            doc = word_app.Documents.Open(
                abs_docx,
                ConfirmConversions=False,
                ReadOnly=True,
                AddToRecentFiles=False,
                Revert=False,
                Format=0,  # wdOpenFormatAuto
                NoEncodingDialog=True
            )
            
            # Ensure pagination is accurate before scanning
            doc.Repaginate()
            try:
                # Force Print Layout for accurate physical page calculations
                if word_app.Windows.Count > 0:
                    word_app.Windows(1).View.Type = 3  # wdPrintView
            except Exception:
                pass
            
            # wdStatisticPages = 2
            page_count = doc.ComputeStatistics(2)
            self.update_status(f"Found {page_count} pages. Starting...", 0.15)

            base_name = os.path.splitext(os.path.basename(abs_docx))[0]
            passwords_data = []

            for i in range(1, page_count + 1):
                # ── Check stop BEFORE doing work ──
                if self.stop_requested:
                    self.update_status(f"Stopped at page {i-1}/{page_count}.", 0)
                    break

                # ── Extract page text and detect identifier ──
                pdf_name = f"{base_name}_page_{i}.pdf"
                detected = "No ID"
                clean_id_for_lookup = "No ID"
                clean_filename = ""

                try:
                    # Robust Page Range Calculation:
                    # wdGoToPage=1, wdGoToAbsolute=1
                    start_pos = doc.GoTo(1, 1, i).Start
                    if i == 1: start_pos = 0 
                    
                    if i < page_count:
                        end_pos = doc.GoTo(1, 1, i + 1).Start
                    else:
                        end_pos = doc.Content.End
                    
                    page_rng = doc.Range(start_pos, end_pos)
                    page_text = page_rng.Text or ""

                    # Normalize text
                    norm_txt = page_text.replace("\r", "\n")
                    pattern = r"(?:n[°ºo]\.?\s*)?(?:d['\u2019´`]\s*)?\s*inscription\s*[:\s]\s*([^\r\n$]+)"
                    match = re.search(pattern, norm_txt, re.IGNORECASE)

                    # Fallback 1: Shapes (Text Boxes)
                    if not match:
                        try:
                            # 3 = wdActiveEndPageNumber
                            for shape in doc.Shapes:
                                try:
                                    if shape.Anchor.Information(3) == i:
                                        if shape.TextFrame.HasText:
                                            txt = shape.TextFrame.TextRange.Text or ""
                                            m = re.search(pattern, txt.replace("\r", "\n"), re.IGNORECASE)
                                            if m: match = m; break
                                except: continue
                        except: pass

                    # Fallback 2: Check ALL Header/Footer types in the current section
                    # Some docs have 'Different First Page' or 'Even/Odd' headers
                    if not match:
                        try:
                            sect = page_rng.Sections(1)
                            for hf in sect.Headers:
                                h_txt = hf.Range.Text or ""
                                m = re.search(pattern, h_txt.replace("\r", "\n"), re.IGNORECASE)
                                if m: match = m; break
                        except: pass

                    # Fallback 3: Content Controls
                    if not match:
                        try:
                            for cc in doc.ContentControls:
                                if cc.Range.InRange(page_rng):
                                    c_txt = cc.Range.Text or ""
                                    m = re.search(pattern, c_txt.replace("\r", "\n"), re.IGNORECASE)
                                    if m: match = m; break
                        except: pass

                    if match:
                        raw_id = match.group(1).strip()
                        id_line = raw_id.split('\n')[0].strip()
                        clean = id_line.replace(" ", "").replace("/", "-")
                        clean = re.sub(r'[\\/:*?"<>|]', '', clean).strip("-")
                        
                        if clean and len(clean) > 0:
                            clean_filename = clean
                            pdf_name = f"{clean}.pdf"
                            detected = f"ID: {clean}"
                            # Normalize for lookup: remove the dashes we added plus spaces/slashes
                            clean_id_for_lookup = re.sub(r'[\s/-]', '', clean)
                except Exception as ex:
                    import traceback
                    traceback.print_exc()
                    print(f"Extraction error on page {i}: {ex}")
                    pass


                # ── Check stop AGAIN before the slow export ──
                if self.stop_requested:
                    self.update_status(f"Stopped at page {i}/{page_count}.", 0)
                    break

                # ── Ensure unique identifier (prevent overwriting) ──
                base_pdf_name = pdf_name
                counter = 1
                pdf_path = os.path.join(os.path.abspath(out_dir), pdf_name)
                while os.path.exists(pdf_path):
                    counter += 1
                    name_part, ext_part = os.path.splitext(base_pdf_name)
                    pdf_name = f"{name_part}_{counter}{ext_part}"
                    pdf_path = os.path.join(os.path.abspath(out_dir), pdf_name)

                # ── Export this single page as PDF ──
                doc.ExportAsFixedFormat(
                    OutputFileName=pdf_path,
                    ExportFormat=17,   # wdExportFormatPDF
                    Range=3,           # wdExportFromTo (3, not 2!)
                    From=i, To=i,
                    Item=0,            # wdExportDocumentContent
                    OpenAfterExport=False,
                    OptimizeFor=0      # wdExportOptimizeForPrint
                )
                
                try:
                    # Determine password: Use mapping if available, otherwise generate random
                    pwd = id_to_password_map.get(clean_id_for_lookup)
                    source_label = "(Excel map)" if (pwd and excel_mapping_path) else "(Randomly generated)"
                    
                    if not pwd:
                        # Fallback: Generate 8-char password (4 numbers, 4 letters)
                        chars = random.choices(string.ascii_letters, k=4) + random.choices(string.digits, k=4)
                        random.shuffle(chars)
                        pwd = ''.join(chars)
                    
                    # Encrypt PDF
                    reader = PdfReader(pdf_path)
                    writer = PdfWriter()
                    for page in reader.pages:
                        writer.add_page(page)
                    writer.encrypt(pwd)
                    with open(pdf_path, "wb") as f:
                        writer.write(f)
                    
                    # Store data for Excel
                    display_id = detected.replace("ID: ", "") if detected != "No ID" else "No ID"
                    passwords_data.append([display_id, pwd, pdf_name])
                except Exception as ex:
                    print(f"Error securing {pdf_name}: {ex}")
                
                prog = 0.15 + (i / page_count) * 0.85
                self.update_status(f"Page {i}/{page_count} → {detected} {source_label}", prog)

            # ── Done ──
            if passwords_data:
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Passwords"
                    ws.append(["ID", "Password", "Filename"])
                    for row in passwords_data:
                        ws.append(row)
                    excel_path = os.path.join(os.path.abspath(out_dir), f"{base_name}_passwords.xlsx")
                    wb.save(excel_path)
                except Exception as ex:
                    print(f"Error saving password Excel file: {ex}")
            
            if not self.stop_requested:
                self.update_status("✓ All pages split successfully!", 1.0)
                self.after(0, self._show_results_btn)
                self.after(100, lambda: messagebox.showinfo(
                    "Complete", f"Split {page_count} pages into PDF files."
                ))
            else:
                self.update_status(f"Stopped by user. Files already saved are in the folder.", 0)
                self.after(0, self._show_results_btn)

        except Exception as e:
            self.update_status("Error occurred.", 0)
            self.after(100, lambda: messagebox.showerror(
                "Failed", f"Conversion failed:\n{str(e)}"
            ))
        finally:
            try:
                if doc:
                    doc.Close(0)
            except Exception:
                pass
            try:
                if word_app:
                    word_app.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
            self.is_processing = False
            self.after(0, self._reset_for_new_run)


if __name__ == "__main__":
    app = SplitterApp()
    app.mainloop()